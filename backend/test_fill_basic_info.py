# -*- coding: utf-8 -*-

from pathlib import Path
import sys
from datetime import datetime, date, time

try:
    import win32com.client as win32
except ImportError:
    print("pywin32가 설치되어 있지 않습니다.")
    print(r".venv\Scripts\python.exe -m pip install pywin32")
    sys.exit(1)

from openpyxl import load_workbook


BASIC_INFO_FILLER_VERSION = "2026-06-30_BASIC_INFO_WITH_TITLE_CHECKBOX"

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

INPUT1_SHEET_NAME = "입력1(학교시설명 등 기본사항 입력) 보고서1,2"
TARGET_TITLE = "학교 환경위생 및 식품위생(정기■ㆍ특별□) 점검표"


def find_latest_file(patterns):
    files = []

    if UPLOAD_DIR.exists():
        for pattern in patterns:
            files.extend(UPLOAD_DIR.glob(pattern))

    if not files:
        return None

    return max(files, key=lambda file: file.stat().st_mtime)


def normalize_value(value):
    if value is None:
        return "-"

    if isinstance(value, datetime):
        return value.strftime("%Y. %m. %d.")

    if isinstance(value, date):
        return value.strftime("%Y. %m. %d.")

    if isinstance(value, time):
        return value.strftime("%H:%M")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    text = str(value).strip()

    if text == "":
        return "-"

    return text


def safe_text(value):
    if value is None:
        return "-"

    text = str(value).strip()

    return text if text else "-"


def find_input1_sheet(workbook):
    if INPUT1_SHEET_NAME in workbook.sheetnames:
        return workbook[INPUT1_SHEET_NAME]

    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("입력1"):
            return workbook[sheet_name]

    return None


def extract_basic_info(excel_path):
    workbook = load_workbook(excel_path, data_only=True, keep_vba=True)
    sheet = find_input1_sheet(workbook)

    if sheet is None:
        raise ValueError(f"입력1 시트를 찾지 못했습니다. 현재 시트: {workbook.sheetnames}")

    return {
        "학교명": normalize_value(sheet["C2"].value),
        "교장": normalize_value(sheet["H2"].value),
        "소재지": normalize_value(sheet["C3"].value),
        "설립구분": normalize_value(sheet["C4"].value),
        "일반교실수": normalize_value(sheet["I4"].value),
        "특별교실수": normalize_value(sheet["K4"].value),
        "전화번호": normalize_value(sheet["C5"].value),
        "FAX번호": normalize_value(sheet["H5"].value),

        "냉방_중앙": normalize_value(sheet["E6"].value),
        "냉방_개별": normalize_value(sheet["H6"].value),
        "난방_중앙": normalize_value(sheet["E7"].value),
        "난방_개별": normalize_value(sheet["H7"].value),
        "환기_중앙": normalize_value(sheet["E8"].value),
        "환기_개별": normalize_value(sheet["H8"].value),

        "먹는물": normalize_value(sheet["J6"].value),
        "저수조": normalize_value(sheet["J7"].value),
        "정수기": normalize_value(sheet["J8"].value),

        "급식시설_조리실": normalize_value(sheet["D9"].value),
        "급식시설_식당": normalize_value(sheet["F9"].value),
        "체육장": normalize_value(sheet["I9"].value),

        "체육관": normalize_value(sheet["D10"].value),
        "강당": normalize_value(sheet["F10"].value),
        "기숙사": normalize_value(sheet["I10"].value),
        "신축년도": normalize_value(sheet["K10"].value),

        "측정일자": normalize_value(sheet["C13"].value),
        "측정시간": normalize_value(sheet["I13"].value),
        "측정장소": normalize_value(sheet["C14"].value),
        "측정자_소속": normalize_value(sheet["D15"].value),
        "측정자_성명": normalize_value(sheet["I15"].value),

        "온도": normalize_value(sheet["C17"].value),
        "습도": normalize_value(sheet["D17"].value),
        "기압": normalize_value(sheet["E17"].value),
        "소음": normalize_value(sheet["F17"].value),
        "CO2": normalize_value(sheet["G17"].value),
        "PM10": normalize_value(sheet["H17"].value),
        "PM2_5": normalize_value(sheet["I17"].value),
        "오존": normalize_value(sheet["J17"].value),
        "환기장치": normalize_value(sheet["K17"].value),
    }


def run(hwp, action_name):
    try:
        hwp.HAction.Run(action_name)
        return True
    except Exception:
        return False


def move_doc_begin(hwp):
    run(hwp, "MoveDocBegin")


def find_text(hwp, text):
    move_doc_begin(hwp)

    try:
        hwp.HAction.GetDefault("RepeatFind", hwp.HParameterSet.HFindReplace.HSet)
        hwp.HParameterSet.HFindReplace.FindString = text
        hwp.HParameterSet.HFindReplace.Direction = hwp.FindDir("Forward")
        hwp.HParameterSet.HFindReplace.IgnoreMessage = 1

        # 일부 한글 버전에서는 아래 속성 설정 시 오류 발생:
        # Property 'HFindReplace.FindType' can not be set.
        # 그래서 사용하지 않음.
        # hwp.HParameterSet.HFindReplace.FindType = 1

        return hwp.HAction.Execute(
            "RepeatFind",
            hwp.HParameterSet.HFindReplace.HSet,
        )

    except Exception as e:
        print(f"[찾기 실패] {text}")
        print(e)
        return False


def all_replace(hwp, find_string, replace_string):
    try:
        hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
        hwp.HParameterSet.HFindReplace.FindString = find_string
        hwp.HParameterSet.HFindReplace.ReplaceString = replace_string
        hwp.HParameterSet.HFindReplace.Direction = hwp.FindDir("AllDoc")
        hwp.HParameterSet.HFindReplace.IgnoreMessage = 1

        # FindType은 한글 버전에 따라 오류가 나므로 사용하지 않음.
        # hwp.HParameterSet.HFindReplace.FindType = 1

        return hwp.HAction.Execute(
            "AllReplace",
            hwp.HParameterSet.HFindReplace.HSet,
        )

    except Exception as e:
        print(f"[제목 치환 실패] {find_string} → {replace_string}")
        print(e)
        return False


def fill_title_checkbox(hwp):
    """
    제목의 정기/특별 체크박스를 정기 체크 상태로 맞춘다.

    목표:
    학교 환경위생 및 식품위생(정기■ㆍ특별□) 점검표
    """

    print()
    print("=" * 60)
    print("[제목 입력] 정기/특별 체크박스 입력")
    print(f"[제목 입력] 목표 제목: {TARGET_TITLE}")
    print("=" * 60)

    if find_text(hwp, TARGET_TITLE):
        print("[제목 입력] 이미 목표 제목 상태입니다.")
        return True

    replacements = [
        # 전체 제목 치환
        ("학교 환경위생 및 식품위생(정기□ㆍ특별□) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기 □ㆍ특별 □) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기□ · 특별□) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기□·특별□) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기▢ㆍ특별▢) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기☐ㆍ특별☐) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기□ㆍ 특별□) 점검표", TARGET_TITLE),
        ("학교 환경위생 및 식품위생(정기 □ㆍ 특별 □) 점검표", TARGET_TITLE),

        # 괄호 안 체크박스만 치환
        ("정기□ㆍ특별□", "정기■ㆍ특별□"),
        ("정기 □ㆍ특별 □", "정기■ㆍ특별□"),
        ("정기□ · 특별□", "정기■ㆍ특별□"),
        ("정기□·특별□", "정기■ㆍ특별□"),
        ("정기▢ㆍ특별▢", "정기■ㆍ특별□"),
        ("정기☐ㆍ특별☐", "정기■ㆍ특별□"),
        ("정기□ㆍ 특별□", "정기■ㆍ특별□"),
        ("정기 □ㆍ 특별 □", "정기■ㆍ특별□"),

        # 특수 기호가 빠져 있는 경우
        ("정기ㆍ특별", "정기■ㆍ특별□"),
        ("정기 ㆍ특별", "정기■ㆍ특별□"),
    ]

    changed_count = 0

    for before, after in replacements:
        if find_text(hwp, before):
            print(f"[제목 입력] 찾음: {before}")

            if all_replace(hwp, before, after):
                changed_count += 1
        else:
            print(f"[제목 입력] 없음: {before}")

    if changed_count > 0:
        print(f"[제목 입력] 완료 / 치환 수: {changed_count}")
        return True

    print("[제목 입력] 정확히 일치하는 제목/체크박스 문구를 찾지 못했습니다.")
    print("[제목 입력] 기본사항 입력은 계속 진행합니다.")
    return False


def insert_text(hwp, text):
    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = safe_text(text)

    return hwp.HAction.Execute(
        "InsertText",
        hwp.HParameterSet.HInsertText.HSet,
    )


def cancel_selection(hwp):
    run(hwp, "Cancel")


def move_cells(hwp, directions):
    action_map = {
        "R": "TableRightCell",
        "L": "TableLeftCell",
        "U": "TableUpperCell",
        "D": "TableLowerCell",
    }

    for direction, count in directions:
        action_name = action_map[direction]

        for _ in range(count):
            if not run(hwp, action_name):
                print(f"[실패] 셀 이동 실패: {direction} {count}")
                return False

    return True


def write_current_cell(hwp, value):
    """
    현재 셀 안의 기존 텍스트를 먼저 지우고 새 값 입력.
    날짜/시간/소속처럼 기존 문구가 뒤에 붙는 문제를 해결하기 위한 함수.
    """

    value = safe_text(value)

    # 1단계: 현재 셀 안의 한 줄 전체 선택 후 삭제
    try:
        run(hwp, "MoveLineBegin")
        run(hwp, "MoveSelLineEnd")
        run(hwp, "Delete")
        cancel_selection(hwp)
    except Exception:
        cancel_selection(hwp)

    # 2단계: 혹시 남아 있으면 한 번 더 삭제 시도
    try:
        run(hwp, "MoveLineBegin")
        run(hwp, "MoveSelLineEnd")
        run(hwp, "Delete")
        cancel_selection(hwp)
    except Exception:
        cancel_selection(hwp)

    # 3단계: 값 입력
    try:
        insert_text(hwp, value)
        return True
    except Exception as e:
        print("[실패] 셀 입력 실패")
        print(e)
        return False


def fill_cell(hwp, label_text, directions, value):
    print(f"검색: {label_text} / 이동: {directions} / 값: {value}")

    found = find_text(hwp, label_text)

    if not found:
        print(f"[실패] '{label_text}'를 찾지 못했습니다.")
        return False

    cancel_selection(hwp)

    if not move_cells(hwp, directions):
        return False

    if write_current_cell(hwp, value):
        print(f"[입력 완료] {label_text} → {value}")
        return True

    return False


def main():
    excel_path = find_latest_file(["*.xlsm", "*.xlsx", "*.xls"])
    hwp_path = find_latest_file(["*.hwp"])

    if excel_path is None:
        print("uploads 폴더에서 엑셀 파일을 찾지 못했습니다.")
        print("웹사이트에서 엑셀 파일을 업로드하고 ANALYZE EXCEL을 먼저 눌러주세요.")
        return

    if hwp_path is None:
        print("uploads 폴더에서 HWP 템플릿을 찾지 못했습니다.")
        print("웹사이트에서 원본 HWP 템플릿을 업로드하고 ANALYZE EXCEL을 먼저 눌러주세요.")
        return

    basic_info = extract_basic_info(excel_path)

    print("=" * 60)
    print(f"BASIC_INFO_FILLER_VERSION: {BASIC_INFO_FILLER_VERSION}")
    print("사용할 엑셀:")
    print(excel_path)
    print()
    print("사용할 한글 템플릿:")
    print(hwp_path)
    print()
    print("엑셀에서 읽은 값:")
    for key, value in basic_info.items():
        print(f"- {key}: {value}")
    print("=" * 60)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    school_name = safe_text(basic_info.get("학교명"))
    output_path = OUTPUT_DIR / f"{timestamp}_{school_name}_기본사항입력테스트.hwp"

    hwp = None

    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")

        try:
            hwp.XHwpWindows.Item(0).Visible = True
        except Exception:
            pass

        try:
            hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        except Exception:
            pass

        hwp.Open(str(hwp_path))

        # 제목/체크박스 먼저 처리
        fill_title_checkbox(hwp)

        jobs = [
            # 기본 개요
            ("학 교 명", [("R", 1)], "학교명"),
            ("교장", [("R", 1)], "교장"),
            ("소 재 지", [("R", 1)], "소재지"),
            ("설립구분", [("R", 1)], "설립구분"),

            # 교실수
            ("교실수", [("R", 2)], "일반교실수"),
            ("교실수", [("R", 4)], "특별교실수"),

            # 연락처
            ("전화번호", [("R", 1)], "전화번호"),
            ("FAX번호", [("R", 1)], "FAX번호"),

            # 설비 현황
            ("냉방", [("R", 2)], "냉방_중앙"),
            ("냉방", [("R", 4)], "냉방_개별"),
            ("난방", [("R", 2)], "난방_중앙"),
            ("난방", [("R", 4)], "난방_개별"),
            ("환기", [("R", 2)], "환기_중앙"),
            ("환기", [("R", 4)], "환기_개별"),

            # 먹는물 시설
            # "먹는물" 직접 검색 금지. "저수조" 기준으로 위쪽 칸에 넣음.
            ("저수조", [("R", 1), ("U", 1)], "먹는물"),
            ("저수조", [("R", 1)], "저수조"),
            ("정수기", [("R", 1)], "정수기"),

            # 급식시설 / 체육장
            ("급식시설", [("R", 2)], "급식시설_조리실"),
            ("급식시설", [("R", 4)], "급식시설_식당"),
            ("체육장", [("R", 1)], "체육장"),

            # 체육관 및 강당
            ("체육관 및 강당", [("R", 2)], "체육관"),
            ("체육관 및 강당", [("R", 4)], "강당"),

            # 기숙사 직접 검색 금지. "신축년도" 기준으로 왼쪽 칸에 넣음.
            ("신축년도", [("L", 1)], "기숙사"),
            ("신축년도", [("R", 1)], "신축년도"),

            # 측정 개요
            ("측정일자", [("R", 1)], "측정일자"),
            ("측정시간", [("R", 1)], "측정시간"),
            ("측정장소", [("R", 1)], "측정장소"),

            # 측정자는 "소속"이 중복될 수 있어서 여기서 한 번만 실행
            ("소속", [("R", 1)], "측정자_소속"),
            ("성명", [("R", 1)], "측정자_성명"),

            # 측정조건
            ("온도", [("D", 1)], "온도"),
            ("습도", [("D", 1)], "습도"),
            ("소음", [("D", 1)], "소음"),
            ("기압", [("D", 1)], "기압"),
            ("CO2", [("D", 1)], "CO2"),
            ("PM10", [("D", 1)], "PM10"),
            ("PM2.5", [("D", 1)], "PM2_5"),
            ("오존", [("D", 1)], "오존"),
        ]

        success_count = 0

        for label_text, directions, key in jobs:
            value = basic_info.get(key, "-")

            if fill_cell(hwp, label_text, directions, value):
                success_count += 1

        hwp.SaveAs(str(output_path))

        print()
        print("[완료] 기본사항 입력 테스트 HWP를 저장했습니다.")
        print(output_path)
        print(f"[결과] 총 {success_count}개 항목 입력 완료")

    except Exception as e:
        print()
        print("[오류] HWP 자동 입력 중 문제가 발생했습니다.")
        print(str(e))

    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass


if __name__ == "__main__":
    main()
    