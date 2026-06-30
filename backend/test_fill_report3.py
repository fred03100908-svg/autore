# -*- coding: utf-8 -*-

from pathlib import Path
from datetime import datetime, date, time

try:
    import win32com.client as win32
except ImportError:
    win32 = None

from openpyxl import load_workbook


REPORT3_FILLER_VERSION = "2026-06-30_REPORT3_STABLE_TOTAL_AIRBORNE_BACTERIA_PATCH"

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

REPORT3_SHEET_NAME = "보고서3"


def find_latest_file(patterns):
    files = []

    if UPLOAD_DIR.exists():
        for pattern in patterns:
            files.extend(UPLOAD_DIR.glob(pattern))

    if not files:
        return None

    return max(files, key=lambda file: file.stat().st_mtime)


def is_empty(value):
    return value is None or str(value).strip() == ""


def decimal_places_from_format(number_format):
    if not number_format:
        return None

    fmt = str(number_format)

    if "." not in fmt:
        if "0" in fmt:
            return 0
        return None

    after_dot = fmt.split(".", 1)[1]
    count = 0

    for ch in after_dot:
        if ch in ["0", "#"]:
            count += 1
        else:
            break

    return count


def format_excel_cell(cell):
    value = cell.value

    if is_empty(value):
        return "-"

    if isinstance(value, datetime):
        return value.strftime("%Y. %m. %d.")

    if isinstance(value, date):
        return value.strftime("%Y. %m. %d.")

    if isinstance(value, time):
        return value.strftime("%H:%M")

    if isinstance(value, str):
        text = value.strip()
        return text if text else "-"

    if isinstance(value, float):
        places = decimal_places_from_format(cell.number_format)

        if places is not None:
            if places == 0:
                return str(int(round(value)))
            return f"{value:.{places}f}"

        if value.is_integer():
            return str(int(value))

        return f"{value:.3f}".rstrip("0").rstrip(".")

    return str(value).strip()


def resize_values(values, row_count):
    values = list(values)

    if len(values) > row_count:
        return values[:row_count]

    while len(values) < row_count:
        values.append("-")

    return values


def read_block(sheet, start_row, end_row, hwp_rows):
    rows = range(start_row, end_row + 1)

    data = {
        "place": [format_excel_cell(sheet.cell(row=r, column=2)) for r in rows],
        "count": [format_excel_cell(sheet.cell(row=r, column=4)) for r in rows],
        "min": [format_excel_cell(sheet.cell(row=r, column=5)) for r in rows],
        "max": [format_excel_cell(sheet.cell(row=r, column=6)) for r in rows],
        "avg": [format_excel_cell(sheet.cell(row=r, column=7)) for r in rows],
        "result": [format_excel_cell(sheet.cell(row=r, column=9)) for r in rows],
    }

    for key in data:
        data[key] = resize_values(data[key], hwp_rows)

    return data


def get_report3_specs():
    return [
        # 3-1
        {"section": "3-1", "name": "PM10", "start_row": 5, "end_row": 11, "hwp_rows": 7},
        {"section": "3-1", "name": "PM2.5", "start_row": 12, "end_row": 16, "hwp_rows": 5},
        {"section": "3-1", "name": "CO2", "start_row": 17, "end_row": 21, "hwp_rows": 5},
        {"section": "3-1", "name": "HCHO", "start_row": 23, "end_row": 30, "hwp_rows": 8},
        {"section": "3-1", "name": "총부유세균", "start_row": 31, "end_row": 34, "hwp_rows": 4},
        {"section": "3-1", "name": "낙하세균", "start_row": 35, "end_row": 38, "hwp_rows": 4},
        {"section": "3-1", "name": "CO", "start_row": 39, "end_row": 41, "hwp_rows": 3},
        {"section": "3-1", "name": "NO2", "start_row": 42, "end_row": 44, "hwp_rows": 3},
        {"section": "3-1", "name": "Rn", "start_row": 45, "end_row": 46, "hwp_rows": 2},
        {"section": "3-1", "name": "석면", "start_row": 47, "end_row": 52, "hwp_rows": 6},
        {"section": "3-1", "name": "오존", "start_row": 53, "end_row": 55, "hwp_rows": 3},
        {"section": "3-1", "name": "진드기", "start_row": 56, "end_row": 56, "hwp_rows": 1},
        {"section": "3-1", "name": "TVOC", "start_row": 57, "end_row": 59, "hwp_rows": 4},
        {"section": "3-1", "name": "벤젠", "start_row": 60, "end_row": 62, "hwp_rows": 4},
        {"section": "3-1", "name": "톨루엔", "start_row": 63, "end_row": 65, "hwp_rows": 4},
        {"section": "3-1", "name": "에틸벤젠", "start_row": 66, "end_row": 68, "hwp_rows": 4},
        {"section": "3-1", "name": "자일렌", "start_row": 69, "end_row": 71, "hwp_rows": 4},
        {"section": "3-1", "name": "스티렌", "start_row": 72, "end_row": 74, "hwp_rows": 4},

        # 3-2
        {"section": "3-2", "name": "조도_칠판면", "start_row": 79, "end_row": 82, "hwp_rows": 4},
        {"section": "3-2", "name": "조도_책상면", "start_row": 83, "end_row": 86, "hwp_rows": 4},
        {"section": "3-2", "name": "조도비_칠판면", "start_row": 87, "end_row": 90, "hwp_rows": 4},
        {"section": "3-2", "name": "조도비_책상면", "start_row": 91, "end_row": 94, "hwp_rows": 4},
        {"section": "3-2", "name": "온도", "start_row": 95, "end_row": 101, "hwp_rows": 7},
        {"section": "3-2", "name": "습도", "start_row": 102, "end_row": 108, "hwp_rows": 7},
        {"section": "3-2", "name": "소음", "start_row": 109, "end_row": 112, "hwp_rows": 4},
        {"section": "3-2", "name": "환기", "start_row": 113, "end_row": 116, "hwp_rows": 4},
    ]


def extract_report3_data(excel_path):
    workbook = load_workbook(excel_path, data_only=True, keep_vba=True)

    if REPORT3_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{REPORT3_SHEET_NAME} 시트를 찾지 못했습니다.")

    sheet = workbook[REPORT3_SHEET_NAME]
    specs = get_report3_specs()

    section_values = {
        "3-1": {
            "place": [],
            "count": [],
            "min": [],
            "max": [],
            "avg": [],
            "result": [],
        },
        "3-2": {
            "place": [],
            "count": [],
            "min": [],
            "max": [],
            "avg": [],
            "result": [],
        },
    }

    blocks = []

    for spec in specs:
        data = read_block(
            sheet=sheet,
            start_row=spec["start_row"],
            end_row=spec["end_row"],
            hwp_rows=spec["hwp_rows"],
        )

        block = dict(spec)
        block["data"] = data
        blocks.append(block)

        section = spec["section"]

        for key in section_values[section]:
            section_values[section][key].extend(data[key])

    return blocks, section_values


def run(hwp, action_name):
    try:
        hwp.HAction.Run(action_name)
        return True
    except Exception:
        return False


def move_doc_begin(hwp):
    run(hwp, "MoveDocBegin")


def find_text(hwp, text, from_begin=True):
    if from_begin:
        move_doc_begin(hwp)

    try:
        hwp.HAction.GetDefault("RepeatFind", hwp.HParameterSet.HFindReplace.HSet)
        hwp.HParameterSet.HFindReplace.FindString = text
        hwp.HParameterSet.HFindReplace.Direction = hwp.FindDir("Forward")
        hwp.HParameterSet.HFindReplace.IgnoreMessage = 1

        return hwp.HAction.Execute(
            "RepeatFind",
            hwp.HParameterSet.HFindReplace.HSet,
        )

    except Exception as e:
        print(f"[찾기 실패] '{text}' 검색 중 오류 발생")
        print(e)
        return False


def insert_text(hwp, text):
    text = "-" if text is None or str(text).strip() == "" else str(text)

    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = text

    return hwp.HAction.Execute(
        "InsertText",
        hwp.HParameterSet.HInsertText.HSet,
    )


def cancel_selection(hwp):
    run(hwp, "Cancel")


def move_right(hwp, count=1):
    for _ in range(count):
        if not run(hwp, "TableRightCell"):
            return False
    return True


def move_left(hwp, count=1):
    for _ in range(count):
        if not run(hwp, "TableLeftCell"):
            return False
    return True


def move_down(hwp, count=1):
    for _ in range(count):
        if not run(hwp, "TableLowerCell"):
            return False
    return True


def clear_current_cell_text(hwp):
    cancel_selection(hwp)

    try:
        run(hwp, "MoveLineBegin")
        run(hwp, "MoveSelLineEnd")
        run(hwp, "Delete")
        cancel_selection(hwp)
        return True
    except Exception:
        cancel_selection(hwp)

    return False


def write_current_cell(hwp, value):
    value = "-" if value is None or str(value).strip() == "" else str(value)

    clear_current_cell_text(hwp)

    try:
        insert_text(hwp, value)
        return True
    except Exception as e:
        print("[실패] 셀 입력 실패")
        print(e)
        return False


def go_to_section(hwp, section_name):
    found = find_text(hwp, section_name, from_begin=True)

    if not found:
        print(f"[실패] HWP에서 '{section_name}'를 찾지 못했습니다.")
        return False

    cancel_selection(hwp)
    return True


def go_to_header_first_data_cell(hwp, section_name, header_candidates):
    for header_text in header_candidates:
        if not go_to_section(hwp, section_name):
            return False

        found = find_text(hwp, header_text, from_begin=False)

        if found:
            cancel_selection(hwp)

            if not move_down(hwp, 1):
                print(f"[실패] '{header_text}' 헤더 아래 첫 데이터 셀로 이동 실패")
                return False

            print(f"[성공] {section_name} 헤더 찾음: {header_text}")
            return True

    print(f"[실패] {section_name}에서 헤더를 찾지 못했습니다. candidates={header_candidates}")
    return False


def fill_column_from_header(hwp, section_name, label, header_candidates, values):
    print()
    print(f"[{section_name}] {label} 열 입력 시작 / 총 {len(values)}행")

    if not go_to_header_first_data_cell(hwp, section_name, header_candidates):
        print(f"[실패] {section_name} / {label} 열 시작 위치 이동 실패")
        return 0

    success_count = 0

    for index, value in enumerate(values, start=1):
        print(f"  {index}/{len(values)} → {value}")

        if write_current_cell(hwp, value):
            success_count += 1

        if index < len(values):
            if not move_down(hwp, 1):
                print(f"[실패] {section_name} / {label} / {index + 1}번째 행 이동 실패")
                return success_count

    return success_count


def write_stable_row_from_current_place(hwp, row_data, label=""):
    """
    현재 위치가 검사장소 칸이라고 가정한다.

    입력하는 칸:
    검사장소 / 검사횟수 / 최소 / 최대 / 평균 / 평가결과

    건드리지 않는 칸:
    검사시간 / 유지기준 / 측정기기 사양 / 검사방법 / 비고
    """

    success_count = 0

    print(f"  {label} / 검사장소 → {row_data.get('place', '-')}")
    if write_current_cell(hwp, row_data.get("place", "-")):
        success_count += 1

    # 검사장소 → 검사시간 건너뜀 → 검사횟수
    if not move_right(hwp, 2):
        print(f"[실패] {label} / 검사횟수 칸 이동 실패")
        return success_count

    print(f"  {label} / 검사횟수 → {row_data.get('count', '-')}")
    if write_current_cell(hwp, row_data.get("count", "-")):
        success_count += 1

    # 검사횟수 → 최소
    if not move_right(hwp, 1):
        print(f"[실패] {label} / 최소 칸 이동 실패")
        return success_count

    print(f"  {label} / 최소 → {row_data.get('min', '-')}")
    if write_current_cell(hwp, row_data.get("min", "-")):
        success_count += 1

    # 최소 → 최대
    if not move_right(hwp, 1):
        print(f"[실패] {label} / 최대 칸 이동 실패")
        return success_count

    print(f"  {label} / 최대 → {row_data.get('max', '-')}")
    if write_current_cell(hwp, row_data.get("max", "-")):
        success_count += 1

    # 최대 → 평균
    if not move_right(hwp, 1):
        print(f"[실패] {label} / 평균 칸 이동 실패")
        return success_count

    print(f"  {label} / 평균 → {row_data.get('avg', '-')}")
    if write_current_cell(hwp, row_data.get("avg", "-")):
        success_count += 1

    # 평균 → 유지기준 건너뜀 → 평가결과
    if not move_right(hwp, 2):
        print(f"[실패] {label} / 평가결과 칸 이동 실패")
        return success_count

    print(f"  {label} / 평가결과 → {row_data.get('result', '-')}")
    if write_current_cell(hwp, row_data.get("result", "-")):
        success_count += 1

    return success_count


def fill_report3_section_3_2_row_based(hwp, values):
    print()
    print("=" * 60)
    print("[3-2] ROW BASED 안정 입력 시작")
    print("[3-2] 검사시간/유지기준은 건드리지 않음")
    print("=" * 60)

    row_count = len(values["place"])

    if not go_to_header_first_data_cell(
        hwp,
        "3-2",
        ["검사장소", "장소"],
    ):
        print("[실패] 3-2 검사장소 첫 데이터 칸으로 이동하지 못했습니다.")
        return 0

    success_count = 0

    for row_index in range(row_count):
        row_data = {
            "place": values["place"][row_index],
            "count": values["count"][row_index],
            "min": values["min"][row_index],
            "max": values["max"][row_index],
            "avg": values["avg"][row_index],
            "result": values["result"][row_index],
        }

        print()
        print(f"[3-2] {row_index + 1}/{row_count}행 입력")

        success_count += write_stable_row_from_current_place(
            hwp,
            row_data,
            label=f"3-2 {row_index + 1}행",
        )

        if row_index < row_count - 1:
            if not move_left(hwp, 7):
                print("[실패] 다음 행 이동 전 검사장소 열 복귀 실패")
                return success_count

            if not move_down(hwp, 1):
                print("[실패] 다음 행 이동 실패")
                return success_count

    print()
    print(f"[3-2] ROW BASED 안정 입력 완료 / 입력 셀 수: {success_count}")

    return success_count


def get_3_2_humidity_values(values):
    humidity_start_index = 4 + 4 + 4 + 4 + 7
    humidity_row_count = 7

    return {
        "place": values["place"][humidity_start_index:humidity_start_index + humidity_row_count],
        "count": values["count"][humidity_start_index:humidity_start_index + humidity_row_count],
        "min": values["min"][humidity_start_index:humidity_start_index + humidity_row_count],
        "max": values["max"][humidity_start_index:humidity_start_index + humidity_row_count],
        "avg": values["avg"][humidity_start_index:humidity_start_index + humidity_row_count],
        "result": values["result"][humidity_start_index:humidity_start_index + humidity_row_count],
    }


def fill_3_2_humidity_only(hwp, values):
    humidity_values = get_3_2_humidity_values(values)
    row_count = len(humidity_values["place"])

    print()
    print("=" * 60)
    print("[3-2 습도 보정] 습도 항목만 다시 입력합니다.")
    print("[3-2 습도 보정] 검사시간/유지기준은 건드리지 않음")
    print("=" * 60)

    if not go_to_section(hwp, "3-2"):
        print("[실패] 3-2 섹션으로 이동 실패")
        return 0

    found = find_text(hwp, "습도", from_begin=False)

    if not found:
        print("[실패] 3-2에서 '습도' 항목을 찾지 못했습니다.")
        return 0

    cancel_selection(hwp)

    if not move_right(hwp, 1):
        print("[실패] 습도 검사장소 첫 칸 이동 실패")
        return 0

    success_count = 0

    for row_index in range(row_count):
        row_data = {
            "place": humidity_values["place"][row_index],
            "count": humidity_values["count"][row_index],
            "min": humidity_values["min"][row_index],
            "max": humidity_values["max"][row_index],
            "avg": humidity_values["avg"][row_index],
            "result": humidity_values["result"][row_index],
        }

        print()
        print(f"[3-2 습도 보정] {row_index + 1}/{row_count}행")

        success_count += write_stable_row_from_current_place(
            hwp,
            row_data,
            label=f"습도 {row_index + 1}행",
        )

        if row_index < row_count - 1:
            if not move_left(hwp, 7):
                print("[실패] 습도 다음 행 이동 전 검사장소 열 복귀 실패")
                return success_count

            if not move_down(hwp, 1):
                print("[실패] 습도 다음 행 이동 실패")
                return success_count

    print()
    print(f"[3-2 습도 보정] 완료 / 입력 셀 수: {success_count}")

    return success_count


def slice_values(values, start_index, row_count):
    return {
        "place": values["place"][start_index:start_index + row_count],
        "count": values["count"][start_index:start_index + row_count],
        "min": values["min"][start_index:start_index + row_count],
        "max": values["max"][start_index:start_index + row_count],
        "avg": values["avg"][start_index:start_index + row_count],
        "result": values["result"][start_index:start_index + row_count],
    }


def normalize_anchor_list(anchor):
    if isinstance(anchor, (list, tuple)):
        return list(anchor)
    return [anchor]


def go_to_item_place_cell(hwp, section_name, item_text, occurrence=1):
    """
    특정 검사항목 셀을 찾고, 그 오른쪽 검사장소 첫 칸으로 이동한다.

    item_text는 문자열 또는 리스트 가능.
    예:
    "낙하세균"
    ["총부유세균", "부유세균"]
    """

    anchors = normalize_anchor_list(item_text)

    for anchor_text in anchors:
        if not go_to_section(hwp, section_name):
            print(f"[실패] {section_name} 섹션 이동 실패")
            return False

        found_all = True

        for index in range(occurrence):
            found = find_text(hwp, anchor_text, from_begin=False)

            if not found:
                found_all = False
                print(f"[검색 실패] {section_name}에서 '{anchor_text}' {index + 1}/{occurrence}번째를 찾지 못했습니다.")
                break

            cancel_selection(hwp)

        if not found_all:
            continue

        if not move_right(hwp, 1):
            print(f"[실패] {section_name} / {anchor_text} 검사장소 칸 이동 실패")
            return False

        print(f"[성공] {section_name} / 기준어 '{anchor_text}'로 검사장소 칸 이동")
        return True

    print(f"[실패] {section_name}에서 기준어 후보 {anchors} 모두 찾지 못했습니다.")
    return False


def fill_block_from_item_anchor(hwp, section_name, item_text, block_values, occurrence=1, label=None):
    row_count = len(block_values["place"])
    success_count = 0
    display_label = label or (item_text[0] if isinstance(item_text, (list, tuple)) else item_text)

    print()
    print("=" * 60)
    print(f"[보정 입력] {section_name} / {display_label} / {row_count}행")
    print("=" * 60)

    if not go_to_item_place_cell(
        hwp=hwp,
        section_name=section_name,
        item_text=item_text,
        occurrence=occurrence,
    ):
        return 0

    for row_index in range(row_count):
        row_data = {
            "place": block_values["place"][row_index],
            "count": block_values["count"][row_index],
            "min": block_values["min"][row_index],
            "max": block_values["max"][row_index],
            "avg": block_values["avg"][row_index],
            "result": block_values["result"][row_index],
        }

        print()
        print(f"[보정 입력] {section_name} / {display_label} / {row_index + 1}/{row_count}행")

        success_count += write_stable_row_from_current_place(
            hwp,
            row_data,
            label=f"{section_name} {display_label} {row_index + 1}행",
        )

        if row_index < row_count - 1:
            if not move_left(hwp, 7):
                print(f"[실패] {section_name} / {display_label} 다음 행 전 검사장소 복귀 실패")
                return success_count

            if not move_down(hwp, 1):
                print(f"[실패] {section_name} / {display_label} 다음 행 이동 실패")
                return success_count

    print()
    print(f"[보정 완료] {section_name} / {display_label} / 입력 셀 수: {success_count}")

    return success_count


def patch_3_1_middle(hwp, values):
    print()
    print("=" * 60)
    print("[3-1 보정] 총부유세균 ~ 오존 구간 보정 입력")
    print("=" * 60)

    patch_specs = [
        # 총부유세균은 HWP에서 줄바꿈 때문에 '총부유세균'으로 안 잡히는 경우가 있어
        # '부유세균'을 fallback anchor로 사용한다.
        {"item": ["총부유세균", "부유세균"], "label": "총부유세균", "start": 25, "rows": 4, "occurrence": 1},
        {"item": "낙하세균", "label": "낙하세균", "start": 29, "rows": 4, "occurrence": 1},
        {"item": "CO", "label": "CO", "start": 33, "rows": 3, "occurrence": 2},
        {"item": "NO2", "label": "NO2", "start": 36, "rows": 3, "occurrence": 1},
        {"item": "Rn", "label": "Rn", "start": 39, "rows": 2, "occurrence": 1},
        {"item": "석면", "label": "석면", "start": 41, "rows": 6, "occurrence": 1},
        {"item": "오존", "label": "오존", "start": 47, "rows": 3, "occurrence": 1},
        {"item": "진드기", "label": "진드기", "start": 50, "rows": 1, "occurrence": 1},
    ]

    total = 0

    for spec in patch_specs:
        block_values = slice_values(
            values=values,
            start_index=spec["start"],
            row_count=spec["rows"],
        )

        total += fill_block_from_item_anchor(
            hwp=hwp,
            section_name="3-1",
            item_text=spec["item"],
            block_values=block_values,
            occurrence=spec["occurrence"],
            label=spec["label"],
        )

    return total


def patch_3_1_voc_tail(hwp, values):
    print()
    print("=" * 60)
    print("[3-1 보정] TVOC ~ 스티렌 구간 보정 입력")
    print("=" * 60)

    patch_specs = [
        {"item": "TVOC", "label": "TVOC", "start": 51, "rows": 4, "occurrence": 1},
        {"item": "벤젠", "label": "벤젠", "start": 55, "rows": 4, "occurrence": 1},
        {"item": "톨루엔", "label": "톨루엔", "start": 59, "rows": 4, "occurrence": 1},
        {"item": "에틸벤젠", "label": "에틸벤젠", "start": 63, "rows": 4, "occurrence": 1},
        {"item": "자일렌", "label": "자일렌", "start": 67, "rows": 4, "occurrence": 1},
        {"item": "스티렌", "label": "스티렌", "start": 71, "rows": 4, "occurrence": 1},
    ]

    total = 0

    for spec in patch_specs:
        block_values = slice_values(
            values=values,
            start_index=spec["start"],
            row_count=spec["rows"],
        )

        total += fill_block_from_item_anchor(
            hwp=hwp,
            section_name="3-1",
            item_text=spec["item"],
            block_values=block_values,
            occurrence=spec["occurrence"],
            label=spec["label"],
        )

    return total


def patch_3_2_noise_ventilation(hwp, values):
    print()
    print("=" * 60)
    print("[3-2 보정] 소음 / 환기 구간 보정 입력")
    print("=" * 60)

    patch_specs = [
        {"item": "소음", "label": "소음", "start": 4 + 4 + 4 + 4 + 7 + 7, "rows": 4, "occurrence": 1},
        {"item": "환기", "label": "환기", "start": 4 + 4 + 4 + 4 + 7 + 7 + 4, "rows": 4, "occurrence": 1},
    ]

    total = 0

    for spec in patch_specs:
        block_values = slice_values(
            values=values,
            start_index=spec["start"],
            row_count=spec["rows"],
        )

        total += fill_block_from_item_anchor(
            hwp=hwp,
            section_name="3-2",
            item_text=spec["item"],
            block_values=block_values,
            occurrence=spec["occurrence"],
            label=spec["label"],
        )

    return total


def fill_report3_section(hwp, section_name, values):
    if section_name == "3-2":
        total = 0

        total += fill_report3_section_3_2_row_based(hwp, values)

        print()
        print("=" * 60)
        print("[3-2] 습도 누락 방지용 보정 입력 시작")
        print("=" * 60)

        total += fill_3_2_humidity_only(hwp, values)

        print()
        print("=" * 60)
        print("[3-2] 소음/환기 누락 방지용 보정 입력 시작")
        print("=" * 60)

        total += patch_3_2_noise_ventilation(hwp, values)

        return total

    total = 0

    total += fill_column_from_header(
        hwp,
        section_name,
        "검사장소",
        ["검사장소", "장소"],
        values["place"],
    )

    total += fill_column_from_header(
        hwp,
        section_name,
        "검사횟수",
        ["검사횟수", "횟수"],
        values["count"],
    )

    total += fill_column_from_header(
        hwp,
        section_name,
        "최소",
        ["최소"],
        values["min"],
    )

    total += fill_column_from_header(
        hwp,
        section_name,
        "최대",
        ["최대"],
        values["max"],
    )

    total += fill_column_from_header(
        hwp,
        section_name,
        "평균",
        ["평균"],
        values["avg"],
    )

    total += fill_column_from_header(
        hwp,
        section_name,
        "평가결과",
        ["평가결과", "평가"],
        values["result"],
    )

    print()
    print("=" * 60)
    print("[3-1] 총부유세균~오존 누락 방지용 보정 입력 시작")
    print("=" * 60)

    total += patch_3_1_middle(hwp, values)

    print()
    print("=" * 60)
    print("[3-1] TVOC~스티렌 누락 방지용 보정 입력 시작")
    print("=" * 60)

    total += patch_3_1_voc_tail(hwp, values)

    return total


def main():
    if win32 is None:
        print("pywin32가 설치되어 있지 않습니다.")
        print(r".venv\Scripts\python.exe -m pip install pywin32")
        return

    excel_path = find_latest_file(["*.xlsm", "*.xlsx", "*.xls"])
    hwp_path = find_latest_file(["*.hwp"])

    if excel_path is None:
        print("uploads 폴더에서 엑셀 파일을 찾지 못했습니다.")
        return

    if hwp_path is None:
        print("uploads 폴더에서 HWP 템플릿을 찾지 못했습니다.")
        return

    blocks, section_values = extract_report3_data(excel_path)

    print("=" * 60)
    print(f"REPORT3_FILLER_VERSION: {REPORT3_FILLER_VERSION}")
    print("사용할 엑셀:")
    print(excel_path)
    print()
    print("사용할 HWP 템플릿:")
    print(hwp_path)
    print()
    print("입력 대상:")

    for block in blocks:
        print(
            f"- {block['section']} / {block['name']} / "
            f"엑셀 {block['start_row']}~{block['end_row']}행 / HWP {block['hwp_rows']}행"
        )

    print()
    print("3-1 총 행 수:", len(section_values["3-1"]["place"]))
    print("3-2 총 행 수:", len(section_values["3-2"]["place"]))
    print("=" * 60)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"{timestamp}_보고서3_TOTAL_AIRBORNE_BACTERIA_PATCH_측정값입력테스트.hwp"

    hwp = None

    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")

        try:
            hwp.XHwpWindows.Item(0).Visible = True
        except Exception:
            pass

        try:
            hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        except Exception:
            pass

        hwp.Open(str(hwp_path))

        total = 0

        print()
        print("=" * 60)
        print("[3-1] 입력 시작")
        print("=" * 60)
        total += fill_report3_section(hwp, "3-1", section_values["3-1"])

        print()
        print("=" * 60)
        print("[3-2] 입력 시작")
        print("=" * 60)
        total += fill_report3_section(hwp, "3-2", section_values["3-2"])

        hwp.SaveAs(str(output_path))

        print()
        print("[완료] 보고서3 안정 입력 테스트 HWP를 저장했습니다.")
        print(output_path)
        print(f"[결과] 총 입력 셀 수: {total}")
        print()
        print("검사시간, 유지기준, 측정기기 사양, 검사방법, 비고는 건드리지 않았습니다.")

    except Exception as e:
        print()
        print("[오류] 보고서3 입력 중 문제가 발생했습니다.")
        print(str(e))

    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass


if __name__ == "__main__":
    main()
    