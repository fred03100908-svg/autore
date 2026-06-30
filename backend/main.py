# -*- coding: utf-8 -*-

from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from openpyxl import load_workbook

import os
import re
import sys
import time
import shutil
import subprocess
import traceback
from datetime import datetime, date, time as dt_time
from pathlib import Path

try:
    import win32com.client as win32
except ImportError:
    win32 = None


MAIN_VERSION = "2026-06-30_FINAL_SERVICE_SAFE_PARSE_EXCEL_REPAIR_BASIC_INFO_THEN_REPORT3"

app = FastAPI()

frontend_origin = os.getenv("FRONTEND_ORIGIN")
frontend_origin_regex = os.getenv(
    "FRONTEND_ORIGIN_REGEX",
    r"^https://.*\.vercel\.app$|^http://localhost:3000$|^http://192\.168\.\d+\.\d+:3000$",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[frontend_origin] if frontend_origin else ["http://localhost:3000"],
    allow_origin_regex=frontend_origin_regex,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

INPUT1_SHEET_NAME = "입력1(학교시설명 등 기본사항 입력) 보고서1,2"


def normalize_value(value):
    """
    JSON 응답으로 안전하게 보낼 수 있는 값으로 변환.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y. %m. %d.")

    if isinstance(value, date):
        return value.strftime("%Y. %m. %d.")

    if isinstance(value, dt_time):
        return value.strftime("%H:%M")

    if isinstance(value, (str, int, float, bool)):
        return value

    return str(value)


def safe_text(value, default=""):
    if value is None:
        return default

    text = str(value).strip()

    if text == "":
        return default

    return text


def sanitize_filename(value):
    text = safe_text(value, "자동생성보고서")
    text = re.sub(r'[\\/:*?"<>|]', "_", text)
    text = text.strip()

    if not text:
        return "자동생성보고서"

    return text


def save_upload_file(upload_file: UploadFile, target_path: Path):
    with open(target_path, "wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)


def clean_folder_files(folder: Path, patterns):
    for pattern in patterns:
        for file in folder.glob(pattern):
            try:
                file.unlink()
            except Exception:
                pass


def find_latest_hwp_after(start_timestamp):
    hwp_files = []

    for file in OUTPUT_DIR.glob("*.hwp"):
        try:
            if file.stat().st_mtime >= start_timestamp:
                hwp_files.append(file)
        except Exception:
            pass

    if not hwp_files:
        return None

    return max(hwp_files, key=lambda file: file.stat().st_mtime)


def find_input1_sheet(workbook):
    if INPUT1_SHEET_NAME in workbook.sheetnames:
        return workbook[INPUT1_SHEET_NAME]

    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("입력1"):
            return workbook[sheet_name]

    return None


def get_cell_value(sheet, cell_address):
    try:
        return normalize_value(sheet[cell_address].value)
    except Exception:
        return None


def repair_excel_for_openpyxl(excel_path: Path) -> Path:
    """
    openpyxl이 스타일 오류로 엑셀을 못 여는 경우,
    Excel COM으로 파일을 열고 .xlsx로 다시 저장해서 복구본을 만든다.

    반환:
    - 복구 성공: repaired_xlsx_path
    - 복구 실패: 원본 excel_path
    """

    if win32 is None:
        print("[EXCEL REPAIR] pywin32가 없어 복구를 건너뜁니다.")
        return excel_path

    repaired_path = excel_path.with_name(
        f"{excel_path.stem}_OPENPYXL_REPAIRED.xlsx"
    )

    excel_app = None
    workbook = None

    try:
        print()
        print("=" * 80)
        print("[EXCEL REPAIR] openpyxl용 엑셀 복구 저장 시작")
        print(f"[EXCEL REPAIR] 원본: {excel_path}")
        print(f"[EXCEL REPAIR] 복구본: {repaired_path}")
        print("=" * 80)
        print()

        excel_app = win32.gencache.EnsureDispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False

        try:
            workbook = excel_app.Workbooks.Open(
                str(excel_path),
                UpdateLinks=0,
                ReadOnly=False,
                CorruptLoad=1,
            )
        except TypeError:
            workbook = excel_app.Workbooks.Open(str(excel_path))

        # 51 = xlsx
        workbook.SaveAs(str(repaired_path), FileFormat=51)
        workbook.Close(SaveChanges=False)
        workbook = None

        now = time.time()
        os.utime(repaired_path, (now, now))

        print("[EXCEL REPAIR] 복구 저장 완료")
        print(repaired_path)

        return repaired_path

    except Exception as e:
        print()
        print("[EXCEL REPAIR 실패] 원본 엑셀을 그대로 사용합니다.")
        print(e)
        return excel_path

    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass

        if excel_app is not None:
            try:
                excel_app.DisplayAlerts = True
                excel_app.Quit()
            except Exception:
                pass


def load_workbook_safe(excel_path: Path):
    """
    openpyxl로 먼저 열어보고,
    실패하면 Excel COM으로 복구 저장한 뒤 다시 연다.

    이번 오류:
    style = self.ws.parent._cell_styles[cell['style_id']]
    IndexError: list index out of range
    """

    try:
        workbook = load_workbook(excel_path, data_only=True, keep_vba=True)
        return workbook, excel_path

    except Exception as e:
        print()
        print("=" * 80)
        print("[OPENPYXL ERROR] 엑셀을 openpyxl로 여는 중 오류 발생")
        print(type(e).__name__)
        print(e)
        print("[OPENPYXL ERROR] Excel COM 복구 저장 후 다시 읽습니다.")
        print("=" * 80)
        print()

        repaired_path = repair_excel_for_openpyxl(excel_path)

        if repaired_path == excel_path:
            raise

        workbook = load_workbook(repaired_path, data_only=True, keep_vba=False)
        return workbook, repaired_path


def sheet_to_matrix(sheet):
    """
    프론트 미리보기용 시트 데이터를 안전하게 2차원 배열로 변환한다.
    모든 행을 동일한 max_column 길이로 맞춘다.
    """

    matrix = []

    max_row = sheet.max_row or 1
    max_column = sheet.max_column or 1

    for row in sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
        values_only=True,
    ):
        safe_row = [normalize_value(cell) for cell in row]

        while len(safe_row) < max_column:
            safe_row.append(None)

        matrix.append(safe_row)

    return matrix


def count_non_empty(matrix):
    count = 0

    for row in matrix:
        for cell in row:
            if cell is not None and str(cell).strip() != "":
                count += 1

    return count


def extract_basic_info(input_sheet):
    """
    입력1 시트 기본정보 추출.
    실제 HWP 입력은 test_fill_basic_info.py가 담당하고,
    여기서는 웹 화면 표시와 다운로드 파일명에 필요한 값만 안전하게 읽는다.
    """

    return {
        "학교명": get_cell_value(input_sheet, "C2"),
        "교장": get_cell_value(input_sheet, "H2"),
        "소재지": get_cell_value(input_sheet, "C3"),
        "설립구분": get_cell_value(input_sheet, "C4"),
        "일반교실수": get_cell_value(input_sheet, "I4"),
        "특별교실수": get_cell_value(input_sheet, "K4"),
        "전화번호": get_cell_value(input_sheet, "C5"),
        "FAX번호": get_cell_value(input_sheet, "H5"),
        "측정일자": get_cell_value(input_sheet, "C13"),
        "측정시간": get_cell_value(input_sheet, "I13"),
        "측정장소": get_cell_value(input_sheet, "C14"),
    }


def read_excel_data(excel_path: Path):
    """
    ANALYZE EXCEL용 안전 분석 함수.
    HWP 실제 입력은 별도 스크립트가 담당한다.
    """

    workbook, excel_path = load_workbook_safe(excel_path)

    input1_sheet = find_input1_sheet(workbook)

    if input1_sheet is None:
        raise ValueError(f"입력1 시트를 찾지 못했습니다. 현재 시트: {workbook.sheetnames}")

    input1_matrix = sheet_to_matrix(input1_sheet)
    basic_info = extract_basic_info(input1_sheet)

    report_sheets = []

    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("보고서"):
            report_sheets.append(workbook[sheet_name])

    report_data = {}

    for sheet in report_sheets:
        try:
            matrix = sheet_to_matrix(sheet)

            report_data[sheet.title] = {
                "sheet_name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "non_empty_count": count_non_empty(matrix),
                "data": matrix,
            }

        except Exception as e:
            report_data[sheet.title] = {
                "sheet_name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "non_empty_count": 0,
                "data": [],
                "error": str(e),
            }

    return {
        "available_sheets": workbook.sheetnames,
        "selected_sheets": {
            "input1": input1_sheet.title,
            "reports": [sheet.title for sheet in report_sheets],
        },
        "basic_info": basic_info,
        "input1_sheet": {
            "sheet_name": input1_sheet.title,
            "max_row": input1_sheet.max_row,
            "max_column": input1_sheet.max_column,
            "non_empty_count": count_non_empty(input1_matrix),
            "data": input1_matrix,
        },
        "report_sheets": report_data,
        "excel_path_used": str(excel_path),
    }


def run_backend_script(script_name: str, timeout_seconds=300):
    script_path = BASE_DIR / script_name

    if not script_path.exists():
        raise FileNotFoundError(f"{script_name} 파일을 찾지 못했습니다: {script_path}")

    print()
    print("=" * 80)
    print(f"[SCRIPT RUN] {script_name}")
    print(f"[PYTHON] {sys.executable}")
    print(f"[CWD] {BASE_DIR}")
    print("=" * 80)
    print()

    result = subprocess.run(
        [sys.executable, str(script_path)],
        cwd=str(BASE_DIR),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout_seconds,
    )

    print()
    print("=" * 80)
    print(f"[{script_name} STDOUT]")
    print(result.stdout)
    print("=" * 80)
    print(f"[{script_name} STDERR]")
    print(result.stderr)
    print("=" * 80)
    print()

    if result.returncode != 0:
        raise RuntimeError(
            f"{script_name} 실행 실패\n"
            f"returncode={result.returncode}\n"
            f"stdout={result.stdout}\n"
            f"stderr={result.stderr}"
        )

    return result


def run_final_hwp_generation():
    """
    최종 서비스 생성 흐름.

    1. test_fill_basic_info.py 실행
       - 제목 체크박스
       - 기본사항 입력

    2. 기본사항이 들어간 HWP를 uploads 폴더에 임시 템플릿으로 복사

    3. test_fill_report3.py 실행
       - 보고서3 입력

    4. 최종 HWP 반환
    """

    print()
    print("=" * 80)
    print(f"[MAIN_VERSION] {MAIN_VERSION}")
    print("[FINAL GENERATE] 기본사항 입력 → 보고서3 입력 순서로 실행합니다.")
    print("=" * 80)
    print()

    basic_start_time = time.time()
    run_backend_script("test_fill_basic_info.py", timeout_seconds=300)

    basic_hwp = find_latest_hwp_after(basic_start_time)

    if basic_hwp is None:
        raise RuntimeError(
            "test_fill_basic_info.py 실행 후 outputs 폴더에서 기본사항 HWP를 찾지 못했습니다."
        )

    print()
    print("=" * 80)
    print("[BASIC OUTPUT FOUND]")
    print(basic_hwp)
    print("=" * 80)
    print()

    temp_template_path = UPLOAD_DIR / "99999999_999999_기본사항완료_보고서3입력용템플릿.hwp"

    for old_temp in UPLOAD_DIR.glob("*기본사항완료_보고서3입력용템플릿.hwp"):
        try:
            old_temp.unlink()
        except Exception:
            pass

    shutil.copyfile(str(basic_hwp), str(temp_template_path))

    now = time.time()
    os.utime(temp_template_path, (now, now))

    print()
    print("=" * 80)
    print("[TEMP TEMPLATE FOR REPORT3]")
    print(temp_template_path)
    print("=" * 80)
    print()

    report3_start_time = time.time()
    run_backend_script("test_fill_report3.py", timeout_seconds=420)

    final_hwp = find_latest_hwp_after(report3_start_time)

    if final_hwp is None:
        raise RuntimeError(
            "test_fill_report3.py 실행 후 outputs 폴더에서 최종 HWP를 찾지 못했습니다."
        )

    print()
    print("=" * 80)
    print("[FINAL HWP FOUND]")
    print(final_hwp)
    print("=" * 80)
    print()

    return final_hwp


@app.get("/")
def root():
    return {
        "message": "AutoReport backend is running",
        "version": MAIN_VERSION,
        "base_dir": str(BASE_DIR),
    }


@app.post("/parse-excel")
async def parse_excel(
    excel: UploadFile = File(...),
    template: UploadFile = File(...),
):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        clean_folder_files(UPLOAD_DIR, ["*.xlsm", "*.xlsx", "*.xls", "*.hwp", "*.hwpx"])

        excel_path = UPLOAD_DIR / f"{timestamp}_{excel.filename}"
        template_path = UPLOAD_DIR / f"{timestamp}_{template.filename}"

        save_upload_file(excel, excel_path)
        save_upload_file(template, template_path)

        result = read_excel_data(excel_path)

        return {
            "ok": True,
            "message": "엑셀과 HWP 템플릿을 업로드하고 엑셀을 읽었습니다.",
            "version": MAIN_VERSION,
            "excel_filename": excel.filename,
            "template_filename": template.filename,
            **result,
        }

    except Exception as e:
        print()
        print("=" * 80)
        print("[PARSE EXCEL ERROR]")
        print(traceback.format_exc())
        print("=" * 80)
        print()

        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "version": MAIN_VERSION,
                "message": f"엑셀 분석 중 오류가 발생했습니다: {str(e)}",
                "traceback": traceback.format_exc(),
            },
        )


@app.post("/generate-report")
async def generate_report(
    excel: UploadFile = File(...),
    template: UploadFile = File(...),
):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        clean_folder_files(UPLOAD_DIR, ["*.xlsm", "*.xlsx", "*.xls", "*.hwp", "*.hwpx"])
        clean_folder_files(OUTPUT_DIR, ["*.hwp", "*.hwpx", "*.txt"])

        excel_path = UPLOAD_DIR / f"{timestamp}_{excel.filename}"
        template_path = UPLOAD_DIR / f"{timestamp}_{template.filename}"

        save_upload_file(excel, excel_path)
        save_upload_file(template, template_path)

        parsed = read_excel_data(excel_path)
        basic_info = parsed.get("basic_info", {})

        final_hwp = run_final_hwp_generation()

        school_name = sanitize_filename(basic_info.get("학교명") or "자동생성보고서")
        download_filename = f"{timestamp}_{school_name}_최종자동생성보고서.hwp"

        return FileResponse(
            path=str(final_hwp),
            filename=download_filename,
            media_type="application/octet-stream",
        )

    except Exception as e:
        print()
        print("=" * 80)
        print("[GENERATE REPORT ERROR]")
        print(traceback.format_exc())
        print("=" * 80)
        print()

        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "version": MAIN_VERSION,
                "message": f"HWP 보고서 생성 중 오류가 발생했습니다: {str(e)}",
                "traceback": traceback.format_exc(),
            },
        )
    