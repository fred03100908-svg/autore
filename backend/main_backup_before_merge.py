from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from openpyxl import load_workbook
import os
import shutil
import traceback
from datetime import datetime, date
from pathlib import Path

try:
    import win32com.client as win32
except ImportError:
    win32 = None

import test_fill_basic_info as basic_filler
import test_fill_report3 as report3_filler


app = FastAPI()

frontend_origin = os.getenv("FRONTEND_ORIGIN")
frontend_origin_regex = os.getenv(
    "FRONTEND_ORIGIN_REGEX",
    r"^https://.*\.vercel\.app$|^http://localhost:3000$",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[frontend_origin] if frontend_origin else ["http://localhost:3000"],
    allow_origin_regex=frontend_origin_regex,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

INPUT1_SHEET_NAME = "입력1(학교시설명 등 기본사항 입력) 보고서1,2"


def normalize_value(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return value


def safe_text(value):
    if value is None:
        return ""
    return str(value)


def sheet_to_matrix(sheet):
    data = []

    for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column,
        values_only=True,
    ):
        data.append([normalize_value(cell) for cell in row])

    return data


def count_non_empty(matrix):
    count = 0

    for row in matrix:
        for cell in row:
            if cell is not None and cell != "":
                count += 1

    return count


def find_input1_sheet(workbook):
    if INPUT1_SHEET_NAME in workbook.sheetnames:
        return workbook[INPUT1_SHEET_NAME]

    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("입력1"):
            return workbook[sheet_name]

    return None


def find_report_sheets(workbook):
    report_sheets = []

    for sheet_name in workbook.sheetnames:
        if sheet_name == "보고서" or sheet_name.startswith("보고서"):
            report_sheets.append(workbook[sheet_name])

    return report_sheets


def extract_basic_info(input_sheet):
    return {
        "학교명": normalize_value(input_sheet["C2"].value),
        "교장": normalize_value(input_sheet["H2"].value),
        "소재지": normalize_value(input_sheet["C3"].value),
        "설립구분": normalize_value(input_sheet["C4"].value),
        "일반교실수": normalize_value(input_sheet["I4"].value),
        "특별교실수": normalize_value(input_sheet["K4"].value),
        "전화번호": normalize_value(input_sheet["C5"].value),
        "FAX번호": normalize_value(input_sheet["H5"].value),

        "냉방_중앙": normalize_value(input_sheet["E6"].value),
        "냉방_개별": normalize_value(input_sheet["H6"].value),
        "난방_중앙": normalize_value(input_sheet["E7"].value),
        "난방_개별": normalize_value(input_sheet["H7"].value),
        "환기_중앙": normalize_value(input_sheet["E8"].value),
        "환기_개별": normalize_value(input_sheet["H8"].value),

        "먹는물": normalize_value(input_sheet["J6"].value),
        "저수조": normalize_value(input_sheet["J7"].value),
        "정수기": normalize_value(input_sheet["J8"].value),

        "급식시설_조리실": normalize_value(input_sheet["D9"].value),
        "급식시설_식당": normalize_value(input_sheet["F9"].value),
        "체육장": normalize_value(input_sheet["I9"].value),

        "체육관": normalize_value(input_sheet["D10"].value),
        "강당": normalize_value(input_sheet["F10"].value),
        "기숙사": normalize_value(input_sheet["I10"].value),
        "신축년도": normalize_value(input_sheet["K10"].value),

        "측정일자": normalize_value(input_sheet["C13"].value),
        "측정시간": normalize_value(input_sheet["I13"].value),
        "측정장소": normalize_value(input_sheet["C14"].value),
        "측정자_소속": normalize_value(input_sheet["D15"].value),
        "측정자_성명": normalize_value(input_sheet["I15"].value),

        "온도": normalize_value(input_sheet["C17"].value),
        "습도": normalize_value(input_sheet["D17"].value),
        "기압": normalize_value(input_sheet["E17"].value),
        "소음": normalize_value(input_sheet["F17"].value),
        "CO2": normalize_value(input_sheet["G17"].value),
        "PM10": normalize_value(input_sheet["H17"].value),
        "PM2_5": normalize_value(input_sheet["I17"].value),
        "오존": normalize_value(input_sheet["J17"].value),
        "환기장치": normalize_value(input_sheet["K17"].value),
    }


def save_upload_file(upload_file: UploadFile, target_path: Path):
    with open(target_path, "wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)


def read_excel_data(excel_path: Path):
    workbook = load_workbook(excel_path, data_only=True, keep_vba=True)

    input1_sheet = find_input1_sheet(workbook)
    report_sheets = find_report_sheets(workbook)

    if input1_sheet is None:
        raise ValueError(f"입력1 시트를 찾지 못했습니다. 현재 시트: {workbook.sheetnames}")

    if len(report_sheets) == 0:
        raise ValueError(f"보고서 시트를 찾지 못했습니다. 현재 시트: {workbook.sheetnames}")

    input1_matrix = sheet_to_matrix(input1_sheet)
    basic_info = extract_basic_info(input1_sheet)

    report_data = {}

    for sheet in report_sheets:
        matrix = sheet_to_matrix(sheet)

        report_data[sheet.title] = {
            "sheet_name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "non_empty_count": count_non_empty(matrix),
            "data": matrix,
        }

    return {
        "available_sheets": workbook.sheetnames,
        "selected_sheets": {
            "input1": input1_sheet.title,
            "reports": [sheet.title for sheet in report_sheets],
        },
        "basic_info": basic_info,
        "input1_sheet": {
            "sheet_name": input1_sheet.title,
            "max_row": input1_sheet.max_row,
            "max_column": input1_sheet.max_column,
            "non_empty_count": count_non_empty(input1_matrix),
            "data": input1_matrix,
        },
        "report_sheets": report_data,
    }


def replace_text_in_hwp(template_path: Path, output_path: Path, mapping: dict):
    if win32 is None:
        raise RuntimeError("pywin32가 설치되어 있지 않습니다. pip install pywin32를 실행하세요.")

    if not template_path.exists():
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")

    hwp = None

    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")

        try:
            hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        except Exception:
            pass

        hwp.Open(str(template_path))

        for key, value in mapping.items():
            find_text = "{{" + key + "}}"
            replace_text = safe_text(value)

            hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
            hwp.HParameterSet.HFindReplace.FindString = find_text
            hwp.HParameterSet.HFindReplace.ReplaceString = replace_text
            hwp.HParameterSet.HFindReplace.Direction = hwp.FindDir("AllDoc")
            hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
            hwp.HParameterSet.HFindReplace.FindType = 1
            hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

        hwp.SaveAs(str(output_path))

    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass


def fill_hwp_with_extracted_data(
    excel_path: Path,
    template_path: Path,
    output_path: Path,
):
    if win32 is None:
        raise RuntimeError("pywin32가 설치되어 있지 않습니다. pip install pywin32를 실행하세요.")

    parsed = read_excel_data(excel_path)
    basic_info = parsed["basic_info"]
    report3_blocks = report3_filler.extract_report3_data(excel_path)

    basic_jobs = [
        ("학 교 명", [("R", 1)], "학교명"),
        ("교장", [("R", 1)], "교장"),
        ("소 재 지", [("R", 1)], "소재지"),
        ("설립구분", [("R", 1)], "설립구분"),
        ("교실수", [("R", 2)], "일반교실수"),
        ("교실수", [("R", 4)], "특별교실수"),
        ("전화번호", [("R", 1)], "전화번호"),
        ("FAX번호", [("R", 1)], "FAX번호"),
        ("냉방", [("R", 2)], "냉방_중앙"),
        ("냉방", [("R", 4)], "냉방_개별"),
        ("난방", [("R", 2)], "난방_중앙"),
        ("난방", [("R", 4)], "난방_개별"),
        ("환기", [("R", 2)], "환기_중앙"),
        ("환기", [("R", 4)], "환기_개별"),
        ("저수조", [("R", 1), ("U", 1)], "먹는물"),
        ("저수조", [("R", 1)], "저수조"),
        ("정수기", [("R", 1)], "정수기"),
        ("급식시설", [("R", 2)], "급식시설_조리실"),
        ("급식시설", [("R", 4)], "급식시설_식당"),
        ("체육장", [("R", 1)], "체육장"),
        ("체육관 및 강당", [("R", 2)], "체육관"),
        ("체육관 및 강당", [("R", 4)], "강당"),
        ("신축년도", [("L", 1)], "기숙사"),
        ("신축년도", [("R", 1)], "신축년도"),
        ("측정일자", [("R", 1)], "측정일자"),
        ("측정시간", [("R", 1)], "측정시간"),
        ("측정장소", [("R", 1)], "측정장소"),
        ("소속", [("R", 1)], "측정자_소속"),
        ("성명", [("R", 1)], "측정자_성명"),
        ("온도", [("D", 1)], "온도"),
        ("습도", [("D", 1)], "습도"),
        ("소음", [("D", 1)], "소음"),
        ("기압", [("D", 1)], "기압"),
        ("CO2", [("D", 1)], "CO2"),
        ("PM10", [("D", 1)], "PM10"),
        ("PM2.5", [("D", 1)], "PM2_5"),
        ("오존", [("D", 1)], "오존"),
    ]

    def make_hwp():
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")

        try:
            hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        except Exception:
            pass

        return hwp

    basic_temp_path = output_path.with_name(f"{output_path.stem}__basic_temp.hwp")
    hwp = None

    try:
        hwp = make_hwp()
        hwp.Open(str(template_path))

        basic_success_count = 0
        for label_text, directions, key in basic_jobs:
            value = basic_info.get(key, "-")
            if basic_filler.fill_cell(hwp, label_text, directions, value):
                basic_success_count += 1

        hwp.SaveAs(str(basic_temp_path))

    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass

    hwp = None

    try:
        hwp = make_hwp()
        hwp.Open(str(basic_temp_path))

        report_success_count = 0
        for block in report3_blocks:
            report_success_count += report3_filler.fill_block(hwp, block)

        hwp.SaveAs(str(output_path))

        return {
            "basic_success_count": basic_success_count,
            "report_success_count": report_success_count,
            "report_blocks": len(report3_blocks),
        }

    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass

        try:
            if basic_temp_path.exists():
                basic_temp_path.unlink()
        except Exception:
            pass


@app.get("/")
def root():
    return {"message": "AutoReport backend is running"}


@app.post("/parse-excel")
async def parse_excel(
    excel: UploadFile = File(...),
    template: UploadFile = File(...),
):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        excel_path = UPLOAD_DIR / f"{timestamp}_{excel.filename}"
        template_path = UPLOAD_DIR / f"{timestamp}_{template.filename}"

        save_upload_file(excel, excel_path)
        save_upload_file(template, template_path)

        result = read_excel_data(excel_path)

        return {
            "ok": True,
            "message": "입력1 시트와 보고서 시트를 성공적으로 읽었습니다.",
            "excel_filename": excel.filename,
            "template_filename": template.filename,
            **result,
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "message": f"엑셀 분석 중 오류가 발생했습니다: {str(e)}",
            },
        )


@app.post("/generate-report")
async def generate_report(
    excel: UploadFile = File(...),
    template: UploadFile = File(...),
):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        excel_path = UPLOAD_DIR / f"{timestamp}_{excel.filename}"
        template_path = UPLOAD_DIR / f"{timestamp}_{template.filename}"

        save_upload_file(excel, excel_path)
        save_upload_file(template, template_path)

        result = read_excel_data(excel_path)
        basic_info = result["basic_info"]

        school_name = safe_text(basic_info.get("학교명")) or "보고서"
        output_filename = f"{timestamp}_{school_name}_자동생성보고서.hwp"
        output_path = OUTPUT_DIR / output_filename

        fill_stats = fill_hwp_with_extracted_data(
            excel_path=excel_path,
            template_path=template_path,
            output_path=output_path,
        )

        return FileResponse(
            path=str(output_path),
            filename=output_filename,
            media_type="application/octet-stream",
        )

    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "message": f"HWP 보고서 생성 중 오류가 발생했습니다: {str(e)}",
            },
        )
    